#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook

from pdf_statement import extract_credit_card_statement_rows


TEXT_JOIN_SEPARATOR = " "
EXCEL_COLUMN_PATTERN = re.compile(r"^[A-Z]+$")
RAW_PREVIEW_ROW_LIMIT = 18
RAW_PREVIEW_COLUMN_LIMIT = 12
TABLE_PREVIEW_ROW_LIMIT = 8
TABLE_PREVIEW_COLUMN_LIMIT = 12
CSV_DELIMITER_CANDIDATES = [",", ";", "\t", "|"]
OPENXML_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
LEGACY_EXCEL_SUFFIXES = {".xls"}
PDF_SUFFIXES = {".pdf"}
STRICT_OPENXML_MAIN_NAMESPACE = b"http://purl.oclc.org/ooxml/spreadsheetml/main"
STRICT_OPENXML_REL_NAMESPACE = b"http://purl.oclc.org/ooxml/officeDocument/relationships"
TRANSITIONAL_OPENXML_MAIN_NAMESPACE = (
    b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
TRANSITIONAL_OPENXML_REL_NAMESPACE = (
    b"http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
SUSPICIOUS_TEXT_MARKERS = ["√", "Ã", "Â", "�"]
DATE_TEXT_PATTERN = re.compile(r"^\d{1,4}[-/]\d{1,2}[-/]\d{1,4}$")
EXCEL_DATE_TOKEN_PATTERN = re.compile(r"(yyyy|yy|mm|m|dd|d)", re.IGNORECASE)


@dataclass
class CanonicalizationResult:
    normalized: pd.DataFrame
    row_count_detected: int
    parse_errors: list[dict[str, Any]]


def camel_to_snake(value: str) -> str:
    return re.sub(r"(?<!^)(?=[A-Z])", "_", value).lower()


def normalize_template_payload(value: Any) -> Any:
    if isinstance(value, dict):
        normalized: dict[str, Any] = {}
        for key, entry in value.items():
            normalized_key = camel_to_snake(key) if isinstance(key, str) else key
            normalized[normalized_key] = normalize_template_payload(entry)
        return normalized
    if isinstance(value, list):
        return [normalize_template_payload(entry) for entry in value]
    return value


def normalize_header(value: Any) -> str:
    text = str(value if value is not None else "").strip()
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    return re.sub(r"\s+", " ", text)


def has_suspicious_text_encoding(value: Any) -> bool:
    text = normalize_text(value)
    return bool(text) and any(marker in text for marker in SUSPICIOUS_TEXT_MARKERS)


def is_blank(value: Any) -> bool:
    return normalize_text(value) == ""


def looks_like_date_header(value: Any) -> bool:
    normalized = normalize_text(value).lower()
    return bool(normalized) and ("fecha" in normalized or "date" in normalized)


def classify_date_representation(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        return "typed_date"

    text = normalize_text(value)
    if not text:
        return None
    if DATE_TEXT_PATTERN.fullmatch(text):
        return "text_date"
    return "other"


def column_letter_to_index(column_letter: str) -> int:
    result = 0
    for character in column_letter:
        result = result * 26 + (ord(character) - ord("A") + 1)
    return result - 1


def index_to_column_letter(index: int) -> str:
    result = ""
    current = index + 1
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(ord("A") + remainder) + result
    return result


def infer_file_kind(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix in PDF_SUFFIXES:
        return "pdf"
    if suffix in LEGACY_EXCEL_SUFFIXES:
        return "xls"
    if suffix in OPENXML_EXCEL_SUFFIXES:
        return "xlsx"
    return "csv"


def is_excel_file_kind(file_kind: str) -> bool:
    return file_kind in {"xls", "xlsx"}


def is_pdf_file_kind(file_kind: str) -> bool:
    return file_kind == "pdf"


def excel_engine_for_kind(file_kind: str) -> str | None:
    if file_kind == "xls":
        return "xlrd"
    return None


def open_excel_file(file_path: Path, *, file_kind: str) -> pd.ExcelFile:
    engine = excel_engine_for_kind(file_kind)
    if engine:
        return pd.ExcelFile(file_path, engine=engine)
    return pd.ExcelFile(file_path)


def read_excel_frame(
    file_path: Path,
    *,
    file_kind: str,
    sheet_name: str | int,
    header: int | None,
    dtype: Any,
    nrows: int | None = None,
) -> pd.DataFrame:
    options: dict[str, Any] = {
        "sheet_name": sheet_name,
        "header": header,
        "dtype": dtype,
    }
    if nrows is not None:
        options["nrows"] = nrows

    engine = excel_engine_for_kind(file_kind)
    if engine:
        options["engine"] = engine

    return pd.read_excel(file_path, **options)


def normalize_strict_openxml_bytes(data: bytes) -> bytes:
    return (
        data.replace(
            STRICT_OPENXML_MAIN_NAMESPACE,
            TRANSITIONAL_OPENXML_MAIN_NAMESPACE,
        ).replace(
            STRICT_OPENXML_REL_NAMESPACE,
            TRANSITIONAL_OPENXML_REL_NAMESPACE,
        )
    )


def is_strict_openxml_workbook(file_path: Path) -> bool:
    try:
        with ZipFile(file_path) as archive:
            workbook_xml = archive.read("xl/workbook.xml")
    except Exception:
        return False

    return (
        STRICT_OPENXML_MAIN_NAMESPACE in workbook_xml
        or STRICT_OPENXML_REL_NAMESPACE in workbook_xml
    )


@contextmanager
def compatible_excel_path(file_path: Path):
    if not is_strict_openxml_workbook(file_path):
        yield file_path
        return

    with TemporaryDirectory() as temp_directory:
        converted_path = Path(temp_directory) / file_path.name
        with ZipFile(file_path) as source_archive, ZipFile(
            converted_path, "w"
        ) as target_archive:
            for entry in source_archive.infolist():
                data = source_archive.read(entry.filename)
                if entry.filename.endswith(".xml") or entry.filename.endswith(".rels"):
                    data = normalize_strict_openxml_bytes(data)
                target_archive.writestr(entry, data)
        yield converted_path


def list_excel_preview_sheet_names(file_path: Path, *, file_kind: str) -> list[str]:
    workbook = open_excel_file(file_path, file_kind=file_kind)
    try:
        if workbook.sheet_names:
            return workbook.sheet_names
    finally:
        workbook.close()

    if file_kind == "xls":
        raise ValueError(
            "The uploaded spreadsheet does not contain any worksheet tabs with rows and columns to preview."
        )

    openpyxl_workbook = load_workbook(file_path, read_only=True, data_only=True)
    non_tabular_sheet_names: list[str] = []
    try:
        worksheet_names = [worksheet.title for worksheet in openpyxl_workbook.worksheets]
        if worksheet_names:
            return worksheet_names
        non_tabular_sheet_names = list(openpyxl_workbook.sheetnames)
    finally:
        openpyxl_workbook.close()

    if non_tabular_sheet_names:
        raise ValueError(
            "The uploaded spreadsheet does not contain any worksheet tabs with rows and columns to preview."
        )

    raise ValueError(
        "The uploaded spreadsheet does not contain any worksheet tabs to preview."
    )


def read_text_with_fallbacks(file_path: Path, encodings: list[str]) -> tuple[str, str]:
    for encoding in encodings:
        try:
            return file_path.read_text(encoding=encoding), encoding
        except UnicodeDecodeError:
            continue
    return file_path.read_text(encoding="latin-1"), "latin-1"


def detect_csv_format(file_path: Path) -> tuple[str, str]:
    sample_text, encoding = read_text_with_fallbacks(
        file_path,
        ["utf-8-sig", "utf-8", "cp1252", "latin-1"],
    )
    sample = sample_text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="".join(CSV_DELIMITER_CANDIDATES))
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    return delimiter, encoding


def load_csv_preview_frame(
    file_path: Path,
    *,
    delimiter: str,
    encoding: str,
) -> pd.DataFrame:
    text, _ = read_text_with_fallbacks(file_path, [encoding])
    rows: list[list[str]] = []
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    for row in reader:
        rows.append([normalize_text(value) for value in row[:RAW_PREVIEW_COLUMN_LIMIT]])
        if len(rows) >= RAW_PREVIEW_ROW_LIMIT:
            break

    width = max((len(row) for row in rows), default=0)
    padded_rows = [row + [""] * (width - len(row)) for row in rows]
    return pd.DataFrame(padded_rows, dtype=object)


def frame_to_coordinate_preview_csv(
    frame: pd.DataFrame,
    *,
    row_offset: int = 0,
    column_offset: int = 0,
) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        ["row"]
        + [index_to_column_letter(column_offset + index) for index in range(len(frame.columns))]
    )
    for index, row in enumerate(frame.itertuples(index=False), start=1):
        writer.writerow(
            [row_offset + index] + [normalize_text(value) for value in row]
        )
    return output.getvalue().strip()


def frame_to_table_preview_csv(frame: pd.DataFrame) -> str:
    output = io.StringIO()
    frame.iloc[:TABLE_PREVIEW_ROW_LIMIT, :TABLE_PREVIEW_COLUMN_LIMIT].to_csv(
        output,
        index=False,
    )
    return output.getvalue().strip()


def load_raw_preview_frame(
    file_path: Path,
    *,
    file_kind: str,
    sheet_name: str | None = None,
    delimiter: str | None = None,
    encoding: str | None = None,
) -> pd.DataFrame:
    if file_kind == "csv":
        frame = load_csv_preview_frame(
            file_path,
            delimiter=delimiter or ",",
            encoding=encoding or "utf-8",
        )
    elif is_excel_file_kind(file_kind):
        frame = read_excel_frame(
            file_path,
            file_kind=file_kind,
            sheet_name=sheet_name or 0,
            header=None,
            dtype=object,
            nrows=RAW_PREVIEW_ROW_LIMIT,
        )
    else:
        raise ValueError(f"Unsupported file kind: {file_kind}")

    return frame.iloc[:, :RAW_PREVIEW_COLUMN_LIMIT].fillna("")


def excel_cell_to_display_value(cell: Any) -> Any:
    value = cell.value
    if value is None:
        return None
    if cell.is_date and isinstance(value, (datetime, date)):
        return format_excel_date_display_value(value, cell.number_format)
    return value


def format_excel_date_display_value(
    value: datetime | date,
    number_format: str | None,
) -> str:
    date_value = value.date() if isinstance(value, datetime) else value
    format_text = str(number_format or "").split(";", 1)[0]
    format_text = format_text.replace("\\-", "-").replace("\\/", "/").strip()
    format_text = re.sub(r'"[^"]*"', "", format_text)

    if not format_text or any(marker in format_text.lower() for marker in ("[$", "[", "]", "h", "s")):
        return date_value.isoformat()

    rendered_parts: list[str] = []
    index = 0
    matched_token = False
    while index < len(format_text):
        match = EXCEL_DATE_TOKEN_PATTERN.match(format_text, index)
        if match:
            token = match.group(1).lower()
            rendered_parts.append(format_excel_date_token(date_value, token))
            index = match.end()
            matched_token = True
            continue

        character = format_text[index]
        if character.isspace():
            index += 1
            continue
        if character in {"/", "-", "."}:
            rendered_parts.append("/")
            index += 1
            continue
        return date_value.isoformat()

    rendered = "".join(rendered_parts).strip("/")
    return rendered if matched_token and rendered else date_value.isoformat()


def format_excel_date_token(value: date, token: str) -> str:
    if token == "yyyy":
        return f"{value.year:04d}"
    if token == "yy":
        return f"{value.year % 100:02d}"
    if token == "mm":
        return f"{value.month:02d}"
    if token == "m":
        return str(value.month)
    if token == "dd":
        return f"{value.day:02d}"
    if token == "d":
        return str(value.day)
    return value.isoformat()


def load_excel_display_frame(
    file_path: Path,
    *,
    file_kind: str,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    if file_kind == "xls":
        frame = read_excel_frame(
            file_path,
            file_kind=file_kind,
            sheet_name=sheet_name or 0,
            header=None,
            dtype=object,
        )
        return frame.fillna("")

    rows: list[list[Any]] = []

    with compatible_excel_path(file_path) as compatible_path:
        workbook = load_workbook(compatible_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
            for row in worksheet.iter_rows():
                rows.append([excel_cell_to_display_value(cell) for cell in row])
        finally:
            workbook.close()

    width = max((len(row) for row in rows), default=0)
    padded_rows = [row + [None] * (width - len(row)) for row in rows]
    return pd.DataFrame(padded_rows, dtype=object)


def build_pdf_preview_csv() -> str:
    frame = pd.DataFrame(
        [
            [
                "PDF statement uploads use ADE first and Gemini fallback during import.",
            ]
        ],
        columns=["document"],
        dtype=object,
    )
    return frame_to_coordinate_preview_csv(frame)


def build_pdf_validation() -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    has_ade_key = bool((os.getenv("ADE_API_KEY") or os.getenv("api_key") or "").strip())
    has_gemini_key = bool((os.getenv("GEMINI_API_KEY") or "").strip())

    if not has_ade_key and not has_gemini_key:
        issues.append(
            {
                "severity": "error",
                "code": "missing_pdf_extractor_credentials",
                "message": "PDF statement imports require ADE_API_KEY or GEMINI_API_KEY for AI extraction.",
                "sheetName": None,
                "columnName": None,
            }
        )
        return {
            "fileKind": "pdf",
            "issues": issues,
        }

    if not has_ade_key:
        issues.append(
            {
                "severity": "warning",
                "code": "pdf_ade_unavailable",
                "message": "ADE_API_KEY is not configured. The importer will fall back to Gemini-only PDF parsing, which is less deterministic.",
                "sheetName": None,
                "columnName": None,
            }
        )
        try:
            from pdf2image import convert_from_path as _convert_from_path
            from PIL import Image as _Image

            _ = _convert_from_path
            _ = _Image
        except Exception:
            issues.append(
                {
                    "severity": "error",
                    "code": "missing_pdf_render_dependencies",
                    "message": "Gemini PDF fallback requires pdf2image and Pillow to render statement pages.",
                    "sheetName": None,
                    "columnName": None,
                }
            )

    return {
        "fileKind": "pdf",
        "issues": issues,
    }


def apply_header_row(
    frame: pd.DataFrame,
    *,
    header_row_index: int,
) -> pd.DataFrame:
    if frame.empty:
        return frame

    header_zero_index = max(header_row_index - 1, 0)
    if header_zero_index >= len(frame.index):
        raise ValueError("Header row index is outside the worksheet bounds.")

    headers = frame.iloc[header_zero_index].tolist()
    data = frame.iloc[header_zero_index + 1 :].reset_index(drop=True)
    data.columns = headers
    return data


def build_workbook_preview(file_path: Path) -> dict[str, Any]:
    file_kind = infer_file_kind(file_path)
    if is_pdf_file_kind(file_kind):
        return {
            "fileKind": file_kind,
            "delimiter": None,
            "encoding": None,
            "sheetPreviews": [
                {
                    "sheetName": "Statement PDF",
                    "previewCsv": build_pdf_preview_csv(),
                }
            ],
        }

    if file_kind == "csv":
        delimiter, encoding = detect_csv_format(file_path)
        frame = load_raw_preview_frame(
            file_path,
            file_kind=file_kind,
            delimiter=delimiter,
            encoding=encoding,
        )
        return {
            "fileKind": file_kind,
            "delimiter": delimiter,
            "encoding": encoding,
            "sheetPreviews": [
                {
                    "sheetName": None,
                    "previewCsv": frame_to_coordinate_preview_csv(frame),
                }
            ],
        }

    sheet_previews = []
    with compatible_excel_path(file_path) as compatible_path:
        for sheet_name in list_excel_preview_sheet_names(
            compatible_path,
            file_kind=file_kind,
        )[:3]:
            frame = load_raw_preview_frame(
                compatible_path,
                file_kind=file_kind,
                sheet_name=sheet_name,
            )
            sheet_previews.append(
                {
                    "sheetName": sheet_name,
                    "previewCsv": frame_to_coordinate_preview_csv(frame),
                }
            )

    return {
        "fileKind": file_kind,
        "delimiter": None,
        "encoding": None,
        "sheetPreviews": sheet_previews,
    }


def build_workbook_validation(file_path: Path) -> dict[str, Any]:
    file_kind = infer_file_kind(file_path)
    if is_pdf_file_kind(file_kind):
        return build_pdf_validation()

    issues: list[dict[str, Any]] = []

    if file_kind == "csv":
        return {
            "fileKind": file_kind,
            "issues": issues,
        }

    strict_openxml = is_strict_openxml_workbook(file_path)
    if strict_openxml:
        issues.append(
            {
                "severity": "warning",
                "code": "strict_openxml_compatibility",
                "message": "The workbook uses Strict Open XML. The importer normalized it for compatibility before parsing.",
                "sheetName": None,
                "columnName": None,
            }
        )

    with compatible_excel_path(file_path) as compatible_path:
        for sheet_name in list_excel_preview_sheet_names(
            compatible_path,
            file_kind=file_kind,
        )[:3]:
            frame = load_raw_preview_frame(
                compatible_path,
                file_kind=file_kind,
                sheet_name=sheet_name,
            )

            if any(has_suspicious_text_encoding(value) for value in frame.to_numpy().flatten().tolist()):
                issues.append(
                    {
                        "severity": "warning",
                        "code": "suspicious_text_encoding",
                        "message": "The sheet contains suspicious character sequences such as '√' or 'Ã', which usually indicates broken text encoding in the source file.",
                        "sheetName": sheet_name,
                        "columnName": None,
                    }
                )

            if frame.empty:
                continue

            header_row = frame.iloc[0].tolist()
            for column_index, header_value in enumerate(header_row):
                if not looks_like_date_header(header_value):
                    continue

                representations = {
                    classification
                    for classification in (
                        classify_date_representation(value)
                        for value in frame.iloc[1:, column_index].tolist()
                    )
                    if classification and classification != "other"
                }
                if "typed_date" in representations and "text_date" in representations:
                    issues.append(
                        {
                            "severity": "warning",
                            "code": "mixed_date_representations",
                            "message": "This date column mixes text dates and Excel date cells. Review imported dates carefully or prefer the original CSV export.",
                            "sheetName": sheet_name,
                            "columnName": normalize_text(header_value) or None,
                        }
                    )

    return {
        "fileKind": file_kind,
        "issues": issues,
    }


def build_table_preview(
    file_path: Path,
    *,
    file_kind: str,
    header_row_index: int,
    rows_to_skip_before_header: int,
    start_column_index: int,
    sheet_name: str | None = None,
    delimiter: str | None = None,
    encoding: str | None = None,
) -> dict[str, Any]:
    header_zero_index = max(header_row_index - 1 - rows_to_skip_before_header, 0)

    if is_pdf_file_kind(file_kind):
        raise ValueError(
            "Table preview is not available for PDF statement imports.",
        )

    if file_kind == "csv":
        frame = pd.read_csv(
            file_path,
            delimiter=delimiter or ",",
            encoding=encoding or "utf-8",
            skiprows=rows_to_skip_before_header,
            header=header_zero_index,
            dtype=object,
        )
    elif is_excel_file_kind(file_kind):
        frame = apply_header_row(
            load_excel_display_frame(
                file_path,
                file_kind=file_kind,
                sheet_name=sheet_name,
            ),
            header_row_index=header_row_index,
        )
    else:
        raise ValueError(f"Unsupported file kind: {file_kind}")

    if start_column_index > 0:
        frame = frame.iloc[:, start_column_index:]
    frame = frame.loc[:, ~frame.columns.map(lambda column: normalize_text(column) == "")]
    frame = frame.fillna("")

    headers = [normalize_text(column) for column in frame.columns[:TABLE_PREVIEW_COLUMN_LIMIT]]
    return {
        "sheetName": sheet_name,
        "previewCsv": frame_to_table_preview_csv(frame),
        "headers": headers,
    }


def resolve_column_name(frame: pd.DataFrame, spec: Any) -> Any | None:
    if spec is None:
        return None

    if isinstance(spec, float) and spec.is_integer():
        spec = int(spec)

    if isinstance(spec, int):
        if 0 <= spec < len(frame.columns):
            return frame.columns[spec]
        if 1 <= spec <= len(frame.columns):
            return frame.columns[spec - 1]
        return None

    candidate = normalize_text(spec)
    if not candidate:
        return None

    if candidate in frame.columns:
        return candidate

    normalized_candidate = normalize_header(candidate)
    header_lookup = {normalize_header(column): column for column in frame.columns}
    if normalized_candidate in header_lookup:
        return header_lookup[normalized_candidate]

    if candidate.isdigit():
        numeric = int(candidate)
        if 1 <= numeric <= len(frame.columns):
            return frame.columns[numeric - 1]
        if 0 <= numeric < len(frame.columns):
            return frame.columns[numeric]

    if EXCEL_COLUMN_PATTERN.fullmatch(candidate.upper()):
        index = column_letter_to_index(candidate.upper())
        if 0 <= index < len(frame.columns):
            return frame.columns[index]

    return None


def resolve_series(frame: pd.DataFrame, spec: Any) -> pd.Series | None:
    if spec is None:
        return None

    if isinstance(spec, list):
        resolved_columns = [resolve_column_name(frame, item) for item in spec]
        resolved_columns = [column for column in resolved_columns if column is not None]
        if not resolved_columns:
            return None
        return frame[resolved_columns].apply(
            lambda row: TEXT_JOIN_SEPARATOR.join(
                [piece for piece in (normalize_text(value) for value in row.tolist()) if piece]
            ).strip(),
            axis=1,
        )

    resolved_name = resolve_column_name(frame, spec)
    if resolved_name is None:
        return None
    return frame[resolved_name]


def decimal_to_string(value: Decimal | None) -> str | None:
    if value is None:
        return None
    normalized = format(value.normalize(), "f")
    if "." in normalized:
        normalized = normalized.rstrip("0").rstrip(".")
    return normalized or "0"


def sanitize_numeric_text(text: str, decimal_hint: str | None, thousands_hint: str | None) -> str:
    clean = text.strip().replace("\u00a0", "").replace(" ", "")
    negative = clean.startswith("(") and clean.endswith(")")
    if negative:
        clean = clean[1:-1]

    clean = re.sub(r"[^\d,.\-+]", "", clean)

    if thousands_hint:
        clean = clean.replace(thousands_hint, "")
    if decimal_hint and decimal_hint != ".":
        clean = clean.replace(decimal_hint, ".")

    if not decimal_hint:
        if "," in clean and "." in clean:
            if clean.rfind(",") > clean.rfind("."):
                clean = clean.replace(".", "").replace(",", ".")
            else:
                clean = clean.replace(",", "")
        elif "," in clean:
            head, tail = clean.rsplit(",", 1)
            clean = f"{head.replace(',', '')}.{tail}" if len(tail) in (1, 2) else clean.replace(",", "")
        elif "." in clean:
            head, tail = clean.rsplit(".", 1)
            clean = f"{head.replace('.', '')}.{tail}" if len(tail) in (1, 2) else clean.replace(".", "")

    if negative and not clean.startswith("-"):
        clean = f"-{clean}"

    return clean


def parse_decimal_value(
    value: Any,
    decimal_hint: str | None = None,
    thousands_hint: str | None = None,
) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        return Decimal(str(value))

    text = normalize_text(value)
    if not text:
        return None

    sanitized = sanitize_numeric_text(text, decimal_hint, thousands_hint)
    if not sanitized or sanitized in {"-", "+", ".", ","}:
        return None

    try:
        return Decimal(sanitized)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid numeric value: {text}") from exc


def parse_date_value(
    value: Any,
    *,
    format_hint: str | None,
    dayfirst: bool,
    reference_date: date,
) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = normalize_text(value)
    if not text:
        return None

    strict_candidates = build_strict_date_candidates(text)
    preferred_candidate = choose_date_candidate(
        strict_candidates,
        dayfirst=dayfirst,
        reference_date=reference_date,
    )
    if preferred_candidate is not None:
        return preferred_candidate.isoformat()

    parsed = pd.NaT
    if format_hint:
        parsed = pd.to_datetime(text, format=format_hint, errors="coerce")
    if pd.isna(parsed):
        parsed = pd.to_datetime(text, dayfirst=dayfirst, errors="coerce")
    if not pd.isna(parsed):
        alternate = pd.to_datetime(text, dayfirst=not dayfirst, errors="coerce")
        if (
            not pd.isna(alternate)
            and parsed.date() > reference_date
            and alternate.date() <= reference_date
        ):
            parsed = alternate
    if pd.isna(parsed):
        raise ValueError(f"Invalid date value: {text}")
    return parsed.date().isoformat()


def parse_reference_date(value: Any) -> date:
    text = normalize_text(value)
    if text:
        try:
            return date.fromisoformat(text)
        except ValueError:
            pass
    return date.today()


def build_strict_date_candidates(value: str) -> dict[bool, date]:
    if not DATE_TEXT_PATTERN.fullmatch(value):
        return {}

    separator = "/" if "/" in value else "-"
    parts = value.split(separator)
    if len(parts) != 3 or not all(part.isdigit() for part in parts):
        return {}

    first, second, third = parts
    candidates: dict[bool, date] = {}

    if len(third) in (2, 4) and len(first) <= 2 and len(second) <= 2:
        year = normalize_candidate_year(third)
        dayfirst_candidate = make_date_candidate(year, int(second), int(first))
        monthfirst_candidate = make_date_candidate(year, int(first), int(second))
    elif len(first) == 4 and len(second) <= 2 and len(third) <= 2:
        year = int(first)
        dayfirst_candidate = make_date_candidate(year, int(third), int(second))
        monthfirst_candidate = make_date_candidate(year, int(second), int(third))
    else:
        return {}

    if dayfirst_candidate is not None:
        candidates[True] = dayfirst_candidate
    if monthfirst_candidate is not None:
        candidates[False] = monthfirst_candidate
    return candidates


def make_date_candidate(year: int, month: int, day: int) -> date | None:
    try:
        return date(year, month, day)
    except ValueError:
        return None


def normalize_candidate_year(value: str) -> int:
    if len(value) == 2:
        return 2000 + int(value)
    return int(value)


def choose_date_candidate(
    candidates: dict[bool, date],
    *,
    dayfirst: bool,
    reference_date: date,
) -> date | None:
    preferred = candidates.get(dayfirst)
    alternate = candidates.get(not dayfirst)
    if preferred is None:
        return alternate
    if alternate is None or alternate == preferred:
        return preferred
    if preferred > reference_date and alternate <= reference_date:
        return alternate
    return preferred


def infer_effective_dayfirst(
    transaction_date_series: pd.Series,
    posted_date_series: pd.Series | None,
    default_dayfirst: bool,
    reference_date: date,
) -> bool:
    scores = {True: 0, False: 0}

    for series in (transaction_date_series, posted_date_series):
        if series is None:
            continue
        for value in series.tolist():
            text = normalize_text(value)
            if not text:
                continue

            candidates = build_strict_date_candidates(text)
            dayfirst_candidate = candidates.get(True)
            monthfirst_candidate = candidates.get(False)

            if (
                dayfirst_candidate is not None
                and monthfirst_candidate is not None
                and dayfirst_candidate != monthfirst_candidate
            ):
                dayfirst_future = dayfirst_candidate > reference_date
                monthfirst_future = monthfirst_candidate > reference_date
                if dayfirst_future != monthfirst_future:
                    scores[False if dayfirst_future else True] += 2
            elif dayfirst_candidate is not None and monthfirst_candidate is None:
                scores[True] += 3
            elif monthfirst_candidate is not None and dayfirst_candidate is None:
                scores[False] += 3

    if scores[True] == scores[False]:
        return default_dayfirst
    return scores[True] > scores[False]


def to_serializable(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return decimal_to_string(value)
    if pd.isna(value):
        return None
    return value


def load_dataframe(file_path: Path, template: dict[str, Any]) -> pd.DataFrame:
    file_kind = str(template.get("file_kind", file_path.suffix.lstrip("."))).lower()
    header_row_index = int(template.get("header_row_index", 1))
    skip_before_header = int(template.get("rows_to_skip_before_header", 0))
    rows_to_skip_after_header = int(template.get("rows_to_skip_after_header", 0))
    normalization_rules = template.get("normalization_rules_json", {}) or {}
    start_column_index = int(normalization_rules.get("start_column_index", 0) or 0)
    header_zero_index = max(header_row_index - 1 - skip_before_header, 0)

    if file_kind == "csv":
        frame = pd.read_csv(
            file_path,
            delimiter=template.get("delimiter", ","),
            encoding=template.get("encoding", "utf-8"),
            skiprows=skip_before_header,
            header=header_zero_index,
            dtype=object,
        )
    elif is_pdf_file_kind(file_kind):
        raise ValueError(
            "PDF statement imports use a dedicated parsing path and do not support dataframe loading.",
        )
    elif is_excel_file_kind(file_kind):
        frame = apply_header_row(
            load_excel_display_frame(
                file_path,
                file_kind=file_kind,
                sheet_name=template.get("sheet_name") or None,
            ),
            header_row_index=header_row_index,
        )
    else:
        raise ValueError(f"Unsupported file kind: {file_kind}")

    if start_column_index > 0:
        frame = frame.iloc[:, start_column_index:]
    if rows_to_skip_after_header > 0:
        frame = frame.iloc[rows_to_skip_after_header:]

    return frame.reset_index(drop=True)


def build_pdf_source_preview_frame(rows: list[dict[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "transaction_date": row.get("transaction_date"),
                "posted_date": row.get("posted_date"),
                "description_raw": row.get("description_raw"),
                "amount_original_signed": row.get("amount_original_signed"),
                "currency_original": row.get("currency_original"),
            }
            for row in rows
        ],
        dtype=object,
    )


def canonicalize_pdf_statement(
    file_path: Path,
    template: dict[str, Any],
    *,
    reference_date: str | None = None,
) -> tuple[pd.DataFrame, CanonicalizationResult]:
    default_currency = (
        normalize_text(template.get("default_currency")).upper() or "EUR"
    )
    reference_date_value = parse_reference_date(reference_date)
    parsed_statement = extract_credit_card_statement_rows(
        str(file_path),
        default_currency=default_currency,
        reference_date=reference_date,
    )

    normalized_rows: list[dict[str, Any]] = []
    parse_errors: list[dict[str, Any]] = []

    for index, row in enumerate(parsed_statement.rows, start=1):
        try:
            transaction_date = parse_date_value(
                row.get("transaction_date"),
                format_hint=None,
                dayfirst=True,
                reference_date=reference_date_value,
            )
            if not transaction_date:
                raise ValueError("Missing transaction date.")

            posted_date = parse_date_value(
                row.get("posted_date"),
                format_hint=None,
                dayfirst=True,
                reference_date=reference_date_value,
            )
            amount = parse_decimal_value(row.get("amount_original_signed"))
            if amount is None:
                raise ValueError("Missing amount.")

            description = normalize_text(row.get("description_raw"))
            if not description:
                raise ValueError("Missing description.")

            balance = parse_decimal_value(row.get("balance_original"))
            raw_row_json = json.dumps(
                {
                    **row,
                    "_source_row": index,
                    "_source_kind": "credit_card_statement_pdf",
                    "_extraction_method": parsed_statement.extraction_method,
                    "_statement_net_total": parsed_statement.statement_net_total,
                    "_parser_model": parsed_statement.parser_model,
                },
                ensure_ascii=False,
                default=str,
            )

            normalized_rows.append(
                {
                    "transaction_date": transaction_date,
                    "posted_date": posted_date,
                    "description_raw": description,
                    "amount_original_signed": decimal_to_string(amount),
                    "currency_original": normalize_text(
                        row.get("currency_original") or default_currency,
                    ).upper()
                    or default_currency,
                    "balance_original": decimal_to_string(balance),
                    "external_reference": normalize_text(
                        row.get("external_reference"),
                    )
                    or None,
                    "transaction_type_raw": normalize_text(
                        row.get("transaction_type_raw"),
                    )
                    or None,
                    "security_symbol": None,
                    "security_name": None,
                    "quantity": None,
                    "unit_price_original": None,
                    "fees_original": None,
                    "fx_rate": None,
                    "raw_row_json": raw_row_json,
                }
            )
        except Exception as exc:
            parse_errors.append({"row": index, "message": str(exc)})

    return (
        build_pdf_source_preview_frame(parsed_statement.rows),
        CanonicalizationResult(
            normalized=pd.DataFrame(normalized_rows, dtype=object),
            row_count_detected=len(parsed_statement.rows),
            parse_errors=parse_errors,
        ),
    )


def load_source_and_canonicalize(
    file_path: Path,
    template: dict[str, Any],
    *,
    reference_date: str | None = None,
) -> tuple[pd.DataFrame, CanonicalizationResult]:
    file_kind = str(template.get("file_kind", file_path.suffix.lstrip("."))).lower()
    if is_pdf_file_kind(file_kind):
        return canonicalize_pdf_statement(
            file_path,
            template,
            reference_date=reference_date,
        )

    frame = load_dataframe(file_path, template)
    canonical = canonicalize_frame(
        frame,
        template,
        reference_date=reference_date,
    )
    return frame, canonical


def resolve_amount_series(
    frame: pd.DataFrame,
    column_map: dict[str, Any],
    sign_logic: dict[str, Any],
    *,
    decimal_hint: str | None,
    thousands_hint: str | None,
) -> pd.Series:
    mode = str(sign_logic.get("mode", "signed_amount"))

    if mode == "debit_credit_columns":
        debit_series = resolve_series(
            frame,
            sign_logic.get("debit_column", column_map.get("debit_amount")),
        )
        credit_series = resolve_series(
            frame,
            sign_logic.get("credit_column", column_map.get("credit_amount")),
        )
        if debit_series is None and credit_series is None:
            raise ValueError("Template sign logic requires debit and/or credit columns.")

        return pd.Series(
            [
                decimal_to_string(
                    (
                        parse_decimal_value(
                            credit,
                            decimal_hint=decimal_hint,
                            thousands_hint=thousands_hint,
                        )
                        or Decimal("0")
                    )
                    - (
                        parse_decimal_value(
                            debit,
                            decimal_hint=decimal_hint,
                            thousands_hint=thousands_hint,
                        )
                        or Decimal("0")
                    )
                )
                for debit, credit in zip(
                    debit_series.tolist() if debit_series is not None else [None] * len(frame),
                    credit_series.tolist() if credit_series is not None else [None] * len(frame),
                    strict=False,
                )
            ]
        )

    amount_series = resolve_series(
        frame,
        sign_logic.get("amount_column", column_map.get("amount_original_signed")),
    )
    if amount_series is None:
        raise ValueError("Template is missing an amount column mapping.")

    if mode == "amount_direction_column":
        direction_series = resolve_series(frame, sign_logic.get("direction_column"))
        if direction_series is None:
            raise ValueError("Template sign logic requires a direction column.")

        debit_values = {
            normalize_text(value).upper()
            for value in (sign_logic.get("debit_values") or ["debit", "out", "sell", "withdrawal"])
        }
        credit_values = {
            normalize_text(value).upper()
            for value in (sign_logic.get("credit_values") or ["credit", "in", "buy", "deposit"])
        }

        def resolve_directional_amount(amount_value: Any, direction_value: Any) -> str | None:
            amount = parse_decimal_value(
                amount_value,
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )
            if amount is None:
                return None
            direction = normalize_text(direction_value).upper()
            if direction in debit_values:
                amount = -abs(amount)
            elif direction in credit_values:
                amount = abs(amount)
            return decimal_to_string(amount)

        return pd.Series(
            [
                resolve_directional_amount(amount_value, direction_value)
                for amount_value, direction_value in zip(
                    amount_series.tolist(),
                    direction_series.tolist(),
                    strict=False,
                )
            ]
        )

    invert_sign = bool(sign_logic.get("invert_sign", False))
    return amount_series.apply(
        lambda value: decimal_to_string(
            -(
                parse_decimal_value(
                    value,
                    decimal_hint=decimal_hint,
                    thousands_hint=thousands_hint,
                )
                or Decimal("0")
            )
            if invert_sign
            else parse_decimal_value(
                value,
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )
        )
    )


def canonicalize_frame(
    frame: pd.DataFrame,
    template: dict[str, Any],
    *,
    reference_date: str | None = None,
) -> CanonicalizationResult:
    column_map = template.get("column_map_json", {}) or {}
    sign_logic = template.get("sign_logic_json", {}) or {}
    normalization_rules = template.get("normalization_rules_json", {}) or {}
    default_currency = str(template.get("default_currency", "EUR")).upper()
    date_format = template.get("date_format")
    decimal_hint = template.get("decimal_separator") or None
    thousands_hint = template.get("thousands_separator") or None
    dayfirst = bool(normalization_rules.get("date_day_first", True))
    reference_date_value = parse_reference_date(reference_date)

    detected_rows = 0
    normalized_rows: list[dict[str, Any]] = []
    parse_errors: list[dict[str, Any]] = []
    source_row_offset = int(template.get("header_row_index", 1)) + int(
        template.get("rows_to_skip_after_header", 0)
    )

    if frame.empty:
        return CanonicalizationResult(
            normalized=pd.DataFrame(normalized_rows),
            row_count_detected=0,
            parse_errors=parse_errors,
        )

    amount_series = resolve_amount_series(
        frame,
        column_map,
        sign_logic,
        decimal_hint=decimal_hint,
        thousands_hint=thousands_hint,
    )
    transaction_date_series = resolve_series(frame, column_map.get("transaction_date"))
    if transaction_date_series is None:
        raise ValueError("Template is missing a transaction_date mapping.")

    description_series = resolve_series(frame, column_map.get("description_raw"))
    posted_date_series = resolve_series(frame, column_map.get("posted_date"))
    currency_series = resolve_series(frame, column_map.get("currency_original"))
    balance_series = resolve_series(frame, column_map.get("balance_original"))
    external_reference_series = resolve_series(frame, column_map.get("external_reference"))
    transaction_type_series = resolve_series(frame, column_map.get("transaction_type_raw"))
    security_symbol_series = resolve_series(frame, column_map.get("security_symbol"))
    security_name_series = resolve_series(frame, column_map.get("security_name"))
    quantity_series = resolve_series(frame, column_map.get("quantity"))
    unit_price_series = resolve_series(frame, column_map.get("unit_price_original"))
    fees_series = resolve_series(frame, column_map.get("fees_original"))
    fx_rate_series = resolve_series(frame, column_map.get("fx_rate"))
    effective_dayfirst = infer_effective_dayfirst(
        transaction_date_series,
        posted_date_series,
        dayfirst,
        reference_date_value,
    )

    for index, (_, row) in enumerate(frame.iterrows()):
        raw_row = {str(column): to_serializable(value) for column, value in row.to_dict().items()}
        if all(is_blank(value) for value in raw_row.values()):
            continue

        detected_rows += 1
        source_row = source_row_offset + index + 1

        try:
            transaction_date = parse_date_value(
                transaction_date_series.iloc[index],
                format_hint=date_format,
                dayfirst=effective_dayfirst,
                reference_date=reference_date_value,
            )
            if not transaction_date:
                raise ValueError("Missing transaction date.")

            posted_date = parse_date_value(
                posted_date_series.iloc[index] if posted_date_series is not None else None,
                format_hint=date_format,
                dayfirst=effective_dayfirst,
                reference_date=reference_date_value,
            )
            amount = parse_decimal_value(
                amount_series.iloc[index],
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )
            if amount is None:
                raise ValueError("Missing amount.")

            description = normalize_text(description_series.iloc[index] if description_series is not None else None)
            transaction_type_raw = normalize_text(
                transaction_type_series.iloc[index] if transaction_type_series is not None else None
            )
            security_symbol = normalize_text(
                security_symbol_series.iloc[index] if security_symbol_series is not None else None
            )
            security_name = normalize_text(
                security_name_series.iloc[index] if security_name_series is not None else None
            )
            external_reference = normalize_text(
                external_reference_series.iloc[index] if external_reference_series is not None else None
            )

            if not description:
                description = TEXT_JOIN_SEPARATOR.join(
                    [piece for piece in [transaction_type_raw, security_symbol, security_name, external_reference] if piece]
                ).strip()
            if not description:
                raise ValueError("Missing description.")

            quantity = parse_decimal_value(
                quantity_series.iloc[index] if quantity_series is not None else None,
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )
            unit_price = parse_decimal_value(
                unit_price_series.iloc[index] if unit_price_series is not None else None,
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )
            balance = parse_decimal_value(
                balance_series.iloc[index] if balance_series is not None else None,
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )
            fees = parse_decimal_value(
                fees_series.iloc[index] if fees_series is not None else None,
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )
            fx_rate = parse_decimal_value(
                fx_rate_series.iloc[index] if fx_rate_series is not None else None,
                decimal_hint=decimal_hint,
                thousands_hint=thousands_hint,
            )

            normalized_rows.append(
                {
                    "transaction_date": transaction_date,
                    "posted_date": posted_date,
                    "description_raw": description,
                    "amount_original_signed": decimal_to_string(amount),
                    "currency_original": normalize_text(
                        currency_series.iloc[index] if currency_series is not None else default_currency
                    ).upper()
                    or default_currency,
                    "balance_original": decimal_to_string(balance),
                    "external_reference": external_reference or None,
                    "transaction_type_raw": transaction_type_raw or None,
                    "security_symbol": security_symbol or None,
                    "security_name": security_name or None,
                    "quantity": decimal_to_string(quantity),
                    "unit_price_original": decimal_to_string(unit_price),
                    "fees_original": decimal_to_string(fees),
                    "fx_rate": decimal_to_string(fx_rate),
                    "raw_row_json": json.dumps(
                        {
                            **raw_row,
                            "_source_row": source_row,
                        },
                        ensure_ascii=False,
                        default=str,
                    ),
                }
            )
        except Exception as exc:
            parse_errors.append({"row": source_row, "message": str(exc)})

    normalized = pd.DataFrame(normalized_rows)
    return CanonicalizationResult(
        normalized=normalized,
        row_count_detected=detected_rows,
        parse_errors=parse_errors,
    )


def build_result(
    mode: str,
    account_id: str,
    template_id: str,
    filename: str,
    frame: pd.DataFrame,
    canonical: CanonicalizationResult,
) -> dict[str, Any]:
    normalized = canonical.normalized
    records = normalized.to_dict(orient="records")
    preview_frame = frame.iloc[:TABLE_PREVIEW_ROW_LIMIT, :TABLE_PREVIEW_COLUMN_LIMIT].fillna("")
    source_table_preview = {
        "headers": [normalize_text(column) for column in preview_frame.columns],
        "rows": [
            {
                str(column): to_serializable(value)
                for column, value in row.items()
            }
            for _, row in preview_frame.iterrows()
        ],
    }
    summary: dict[str, Any] = {
        "schemaVersion": "v1",
        "accountId": account_id,
        "templateId": template_id,
        "originalFilename": filename,
        "rowCountDetected": canonical.row_count_detected,
        "rowCountParsed": len(records),
        "rowCountDuplicates": 0,
        "rowCountFailed": len(canonical.parse_errors),
        "dateRange": {
            "start": min((row["transaction_date"] for row in records), default=None),
            "end": max((row["transaction_date"] for row in records), default=None),
        }
        if records
        else None,
        "sourceTablePreview": source_table_preview,
        "normalizedRows": records,
        "sampleRows": records[:5],
        "parseErrors": canonical.parse_errors,
    }
    if mode == "commit":
        summary["importBatchId"] = str(uuid.uuid4())
        summary["rowCountInserted"] = len(records)
        summary["transactionIds"] = [str(uuid.uuid4()) for _ in records]
        summary["jobsQueued"] = [
            "classification",
            "transfer_rematch",
            "position_rebuild",
            "metric_refresh",
        ]
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Template-driven pandas ingest wrapper")
    parser.add_argument(
        "mode",
        choices=[
            "preview",
            "commit",
            "inspect-workbook",
            "preview-table",
            "validate-workbook",
        ],
    )
    parser.add_argument("--file-path", required=True)
    parser.add_argument("--account-id")
    parser.add_argument("--template-id")
    parser.add_argument("--template-json")
    parser.add_argument("--file-kind")
    parser.add_argument("--sheet-name")
    parser.add_argument("--header-row-index", type=int)
    parser.add_argument("--rows-to-skip-before-header", type=int, default=0)
    parser.add_argument("--start-column-index", type=int, default=0)
    parser.add_argument("--delimiter")
    parser.add_argument("--encoding")
    parser.add_argument("--reference-date")
    args = parser.parse_args()

    file_path = Path(args.file_path)
    if args.mode == "inspect-workbook":
        result = build_workbook_preview(file_path)
        sys.stdout.write(json.dumps(result, ensure_ascii=False))
        return 0

    if args.mode == "validate-workbook":
        result = build_workbook_validation(file_path)
        sys.stdout.write(json.dumps(result, ensure_ascii=False))
        return 0

    if args.mode == "preview-table":
        if not args.header_row_index:
            raise ValueError("--header-row-index is required for preview-table.")
        result = build_table_preview(
            file_path,
            file_kind=(args.file_kind or infer_file_kind(file_path)).lower(),
            sheet_name=args.sheet_name or None,
            header_row_index=args.header_row_index,
            rows_to_skip_before_header=args.rows_to_skip_before_header,
            start_column_index=args.start_column_index,
            delimiter=args.delimiter,
            encoding=args.encoding,
        )
        sys.stdout.write(json.dumps(result, ensure_ascii=False))
        return 0

    if not args.account_id or not args.template_id or not args.template_json:
        raise ValueError(
            "--account-id, --template-id, and --template-json are required for preview and commit.",
        )

    template = normalize_template_payload(json.loads(args.template_json))
    frame, canonical = load_source_and_canonicalize(
        file_path,
        template,
        reference_date=args.reference_date,
    )
    result = build_result(
        args.mode,
        args.account_id,
        args.template_id,
        file_path.name,
        frame,
        canonical,
    )
    sys.stdout.write(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        sys.stderr.write(f"{exc}\n")
        raise SystemExit(1)
