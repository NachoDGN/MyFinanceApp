#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ID_NAMESPACE = uuid.UUID("6d34665c-a087-4cc4-986c-d18865dcf72a")


def parse_args() -> argparse.Namespace:
  parser = argparse.ArgumentParser(
    description=(
      "Generate a Supabase migration that imports a workbook of historical "
      "fund NAV prices into public.securities, public.security_aliases, and "
      "public.security_prices."
    ),
  )
  parser.add_argument("--workbook", required=True, type=Path)
  parser.add_argument("--output", required=True, type=Path)
  parser.add_argument("--name", required=True)
  parser.add_argument("--isin")
  parser.add_argument("--provider-name", default="manual_fund_nav")
  parser.add_argument("--provider-symbol")
  parser.add_argument("--canonical-symbol")
  parser.add_argument("--display-symbol")
  parser.add_argument("--exchange-name", default="MANUAL")
  parser.add_argument("--asset-type", default="other")
  parser.add_argument("--quote-currency", default="EUR")
  parser.add_argument("--country")
  parser.add_argument("--mic-code")
  parser.add_argument("--figi")
  parser.add_argument("--price-source", default="manual_nav_import")
  parser.add_argument("--market-state", default="official_nav")
  parser.add_argument("--sheet-name")
  parser.add_argument("--date-column", default="Date")
  parser.add_argument("--close-column", default="Close")
  parser.add_argument("--date-format", default="%A, %B %d, %Y")
  parser.add_argument("--alias", action="append", default=[])
  parser.add_argument("--metadata-json", default="{}")
  parser.add_argument(
    "--merge-metadata-json-on-conflict",
    action="store_true",
    help=(
      "Merge incoming metadata_json into existing security metadata on upsert "
      "instead of replacing it outright."
    ),
  )
  parser.add_argument(
    "--preserve-existing-security-fields-on-conflict",
    action="store_true",
    help=(
      "Keep existing security identity fields on conflict and update only "
      "metadata/refresh fields."
    ),
  )
  return parser.parse_args()


def sql_string(value: str | None) -> str:
  if value is None:
    return "null"
  escaped = value.replace("'", "''")
  return f"'{escaped}'"


def sql_bool(value: bool) -> str:
  return "true" if value else "false"


def sql_json(value: Any) -> str:
  return sql_string(json.dumps(value, sort_keys=True, separators=(",", ":")))


def format_decimal(value: Decimal) -> str:
  normalized = value.normalize()
  text = format(normalized, "f")
  if "." in text:
    text = text.rstrip("0").rstrip(".")
  return text or "0"


def normalize_alias_text(value: str) -> str:
  return " ".join(value.strip().upper().split())


def parse_date_cell(value: object, date_format: str) -> str:
  if isinstance(value, datetime):
    return value.date().isoformat()
  if isinstance(value, date):
    return value.isoformat()
  if isinstance(value, str):
    return datetime.strptime(value.strip(), date_format).date().isoformat()
  raise ValueError(f"Unsupported date cell value: {value!r}")


def parse_price_cell(value: object) -> Decimal:
  try:
    price = Decimal(str(value))
  except (InvalidOperation, TypeError) as exc:
    raise ValueError(f"Unsupported price cell value: {value!r}") from exc
  if price <= 0:
    raise ValueError(f"Price must be positive, received {price}")
  return price


def load_price_rows(
  workbook_path: Path,
  sheet_name: str | None,
  date_column: str,
  close_column: str,
  date_format: str,
) -> tuple[str, list[tuple[str, Decimal]]]:
  workbook = load_workbook(workbook_path, data_only=True, read_only=True)
  worksheet = workbook[sheet_name] if sheet_name else workbook.active
  header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
  header_index = {
    str(value).strip(): index for index, value in enumerate(header_row) if value is not None
  }

  if date_column not in header_index:
    raise ValueError(
      f"Column {date_column!r} was not found in workbook header {list(header_index)}."
    )
  if close_column not in header_index:
    raise ValueError(
      f"Column {close_column!r} was not found in workbook header {list(header_index)}."
    )

  date_index = header_index[date_column]
  close_index = header_index[close_column]
  rows: list[tuple[str, Decimal]] = []

  for row_number, row in enumerate(
    worksheet.iter_rows(min_row=2, values_only=True),
    start=2,
  ):
    raw_date = row[date_index] if date_index < len(row) else None
    raw_close = row[close_index] if close_index < len(row) else None
    if raw_date in (None, "") and raw_close in (None, ""):
      continue
    if raw_date in (None, ""):
      raise ValueError(f"Row {row_number} is missing a date value.")

    try:
      parsed_date = parse_date_cell(raw_date, date_format)
    except ValueError:
      if raw_close in (None, ""):
        continue
      raise ValueError(
        f"Row {row_number} has an unparseable date value: {raw_date!r}."
      ) from None

    if raw_close in (None, ""):
      raise ValueError(f"Row {row_number} is missing a close value.")

    rows.append((parsed_date, parse_price_cell(raw_close)))

  rows.sort(key=lambda item: item[0])
  if not rows:
    raise ValueError("Workbook did not contain any data rows.")

  deduped_rows: list[tuple[str, Decimal]] = []
  seen_by_date: dict[str, Decimal] = {}
  for parsed_date, parsed_price in rows:
    existing_price = seen_by_date.get(parsed_date)
    if existing_price is None:
      seen_by_date[parsed_date] = parsed_price
      deduped_rows.append((parsed_date, parsed_price))
      continue
    if existing_price != parsed_price:
      raise ValueError(
        f"Workbook contains conflicting prices for {parsed_date}: "
        f"{existing_price} vs {parsed_price}."
      )

  return worksheet.title, deduped_rows


def stable_uuid(label: str) -> str:
  return str(uuid.uuid5(ID_NAMESPACE, label))


def build_migration_sql(
  *,
  workbook_path: Path,
  source_label: str,
  security_id: str,
  name: str,
  provider_name: str,
  provider_symbol: str,
  canonical_symbol: str,
  display_symbol: str,
  exchange_name: str,
  asset_type: str,
  quote_currency: str,
  country: str | None,
  mic_code: str | None,
  isin: str | None,
  figi: str | None,
  price_source: str,
  market_state: str,
  metadata_json: dict[str, Any],
  sheet_name: str | None,
  aliases: list[str],
  price_rows: list[tuple[str, Decimal]],
  price_raw_json: dict[str, Any] | None = None,
  merge_metadata_json_on_conflict: bool = False,
  preserve_existing_security_fields_on_conflict: bool = False,
) -> str:
  latest_price_date = price_rows[-1][0]
  last_price_refresh_at = f"{latest_price_date}T16:00:00Z"
  security_lookup_sql = (
    f"(select id from public.securities where provider_name = {sql_string(provider_name)} "
    f"and provider_symbol = {sql_string(provider_symbol)} limit 1)"
  )
  if price_raw_json is None:
    price_raw_json = {
      "importSource": "workbook",
      "sourceWorkbook": workbook_path.name,
      "sheetName": sheet_name,
      "priceType": "nav",
    }
  metadata_assignment_sql = (
    "coalesce(public.securities.metadata_json, '{}'::jsonb) || excluded.metadata_json"
    if merge_metadata_json_on_conflict
    else "excluded.metadata_json"
  )
  security_field_assignment = (
    {
      "canonical_symbol": "public.securities.canonical_symbol",
      "display_symbol": "public.securities.display_symbol",
      "name": "public.securities.name",
      "exchange_name": "public.securities.exchange_name",
      "mic_code": "public.securities.mic_code",
      "asset_type": "public.securities.asset_type",
      "quote_currency": "public.securities.quote_currency",
      "country": "public.securities.country",
      "isin": "public.securities.isin",
      "figi": "public.securities.figi",
    }
    if preserve_existing_security_fields_on_conflict
    else {
      "canonical_symbol": "excluded.canonical_symbol",
      "display_symbol": "excluded.display_symbol",
      "name": "excluded.name",
      "exchange_name": "excluded.exchange_name",
      "mic_code": "excluded.mic_code",
      "asset_type": "excluded.asset_type",
      "quote_currency": "excluded.quote_currency",
      "country": "excluded.country",
      "isin": "excluded.isin",
      "figi": "excluded.figi",
    }
  )

  price_values_sql = ",\n".join(
    [
      "  ("
      + ", ".join(
        [
          sql_string(price_date),
          sql_string(f"{price_date}T16:00:00Z"),
          format_decimal(price),
          sql_string(quote_currency),
          sql_string(price_source),
          sql_bool(False),
          sql_bool(True),
          sql_string(market_state),
          sql_json(price_raw_json),
        ]
      )
      + ")"
      for price_date, price in price_rows
    ]
  )

  alias_rows = list(
    dict.fromkeys(
      normalize_alias_text(alias)
      for alias in aliases
      if alias and alias.strip()
    )
  )
  alias_sql = ""
  if alias_rows:
    alias_values_sql = ",\n".join(
      [
        "  ("
        + ", ".join(
          [
            sql_string(stable_uuid(f"security-alias:{provider_name}:{provider_symbol}:{alias}")),
            sql_string(alias),
            sql_string("manual"),
            "1.0",
          ]
        )
        + ")"
        for alias in alias_rows
      ]
    )
    alias_sql = f"""
insert into public.security_aliases (
  id,
  security_id,
  alias_text_normalized,
  alias_source,
  confidence
)
select
  alias_data.id::uuid,
  security.id,
  alias_data.alias_text_normalized,
  alias_data.alias_source,
  alias_data.confidence::numeric
from {security_lookup_sql} as security(id)
cross join (
values
{alias_values_sql}
) as alias_data(id, alias_text_normalized, alias_source, confidence)
on conflict (security_id, alias_text_normalized) do nothing;
"""

  return f"""-- Generated by scripts/generate_fund_nav_migration.py from {source_label}

insert into public.securities (
  id,
  provider_name,
  provider_symbol,
  canonical_symbol,
  display_symbol,
  name,
  exchange_name,
  mic_code,
  asset_type,
  quote_currency,
  country,
  isin,
  figi,
  active,
  metadata_json,
  last_price_refresh_at
)
values (
  {sql_string(security_id)}::uuid,
  {sql_string(provider_name)},
  {sql_string(provider_symbol)},
  {sql_string(canonical_symbol)},
  {sql_string(display_symbol)},
  {sql_string(name)},
  {sql_string(exchange_name)},
  {sql_string(mic_code)},
  {sql_string(asset_type)},
  {sql_string(quote_currency)},
  {sql_string(country)},
  {sql_string(isin)},
  {sql_string(figi)},
  true,
  {sql_json(metadata_json)}::jsonb,
  {sql_string(last_price_refresh_at)}::timestamptz
)
on conflict (provider_name, provider_symbol) do update
set
  canonical_symbol = {security_field_assignment["canonical_symbol"]},
  display_symbol = {security_field_assignment["display_symbol"]},
  name = {security_field_assignment["name"]},
  exchange_name = {security_field_assignment["exchange_name"]},
  mic_code = {security_field_assignment["mic_code"]},
  asset_type = {security_field_assignment["asset_type"]},
  quote_currency = {security_field_assignment["quote_currency"]},
  country = {security_field_assignment["country"]},
  isin = {security_field_assignment["isin"]},
  figi = {security_field_assignment["figi"]},
  active = excluded.active,
  metadata_json = {metadata_assignment_sql},
  last_price_refresh_at = excluded.last_price_refresh_at;
{alias_sql}
insert into public.security_prices (
  security_id,
  price_date,
  quote_timestamp,
  price,
  currency,
  source_name,
  is_realtime,
  is_delayed,
  market_state,
  raw_json
)
select
  security.id,
  price_data.price_date::date,
  price_data.quote_timestamp::timestamptz,
  price_data.price::numeric,
  price_data.currency,
  price_data.source_name,
  price_data.is_realtime,
  price_data.is_delayed,
  price_data.market_state,
  price_data.raw_json::jsonb
from {security_lookup_sql} as security(id)
cross join (
values
{price_values_sql}
) as price_data(
  price_date,
  quote_timestamp,
  price,
  currency,
  source_name,
  is_realtime,
  is_delayed,
  market_state,
  raw_json
)
on conflict (security_id, price_date, source_name)
do update set
  quote_timestamp = excluded.quote_timestamp,
  price = excluded.price,
  currency = excluded.currency,
  is_realtime = excluded.is_realtime,
  is_delayed = excluded.is_delayed,
  market_state = excluded.market_state,
  raw_json = excluded.raw_json;
"""


def main() -> None:
  args = parse_args()
  workbook_path = args.workbook.expanduser().resolve()
  if not workbook_path.exists():
    raise FileNotFoundError(f"Workbook not found: {workbook_path}")

  provider_symbol = args.provider_symbol or args.isin
  if not provider_symbol:
    raise ValueError("Either --provider-symbol or --isin must be provided.")

  metadata_json = json.loads(args.metadata_json)
  if not isinstance(metadata_json, dict):
    raise ValueError("--metadata-json must decode to a JSON object.")

  sheet_name, price_rows = load_price_rows(
    workbook_path,
    args.sheet_name,
    args.date_column,
    args.close_column,
    args.date_format,
  )

  security_id_key = args.isin or f"{args.provider_name}:{provider_symbol}"
  security_id = stable_uuid(f"security:{security_id_key}")
  canonical_symbol = args.canonical_symbol or args.display_symbol or provider_symbol
  display_symbol = args.display_symbol or canonical_symbol

  sql_text = build_migration_sql(
    workbook_path=workbook_path,
    source_label=workbook_path.name,
    security_id=security_id,
    name=args.name,
    provider_name=args.provider_name,
    provider_symbol=provider_symbol,
    canonical_symbol=canonical_symbol,
    display_symbol=display_symbol,
    exchange_name=args.exchange_name,
    asset_type=args.asset_type,
    quote_currency=args.quote_currency,
    country=args.country,
    mic_code=args.mic_code,
    isin=args.isin,
    figi=args.figi,
    price_source=args.price_source,
    market_state=args.market_state,
    metadata_json=metadata_json,
    sheet_name=sheet_name,
    aliases=args.alias,
    price_rows=price_rows,
    merge_metadata_json_on_conflict=args.merge_metadata_json_on_conflict,
    preserve_existing_security_fields_on_conflict=(
      args.preserve_existing_security_fields_on_conflict
    ),
  )

  output_path = args.output.expanduser().resolve()
  output_path.parent.mkdir(parents=True, exist_ok=True)
  output_path.write_text(sql_text, encoding="utf-8")

  print(
    json.dumps(
      {
        "output": str(output_path),
        "securityId": security_id,
        "providerName": args.provider_name,
        "providerSymbol": provider_symbol,
        "sheetName": sheet_name,
        "rows": len(price_rows),
        "dateRange": {
          "start": price_rows[0][0],
          "end": price_rows[-1][0],
        },
      },
      indent=2,
    )
  )


if __name__ == "__main__":
  main()
